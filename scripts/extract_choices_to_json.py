#This file takes an XLSForm from a Survey123 survey, reads the relevant soil and hydrology indicators, creates a dictionary to store them, and then outputs that dictionary into a JSON file at output/great_plains/review/indicator_mapping.json
#File paths will be changed depending on what survey we are working with


#Rename the XLSForm to {CURRENT_SURVEY}.xlsx

from openpyxl import load_workbook
from pathlib import Path
import json

#Type the name of the survey you are currently parsing (ie: great_plains, arid_west, etc.)
CURRENT_SURVEY = "great_plains"

XLSFORM_PATH = Path(f"data/{CURRENT_SURVEY}/xlsform/{CURRENT_SURVEY}.xlsx")
OUTPUT_PATH = Path(f"output/{CURRENT_SURVEY}/review/{CURRENT_SURVEY}_indicator_mapping.json")

CATEGORY_MAP = {
    #Hydric soil indicators
    "hyd_soil_indicators_a": "Hydric Soil",
    "hyd_soil_indicators_s": "Hydric Soil",
    "hyd_soil_indicators_f": "Hydric Soil",

    #Problematic Hydric Soil Indicators
    "prb_hyd_soil_indicators_a": "Hydric Soil",
    "prb_hyd_soil_indicators_f": "Hydric Soil",
    "prb_hyd_soil_indicators_other": "Hydric Soil",

    #Hydrology indicators
    "prim_hydro_wetland_a": "Hydrology",
    "prim_hydro_wetland_b": "Hydrology",
    "prim_hydro_wetland_c": "Hydrology",
    "prim_hydro_wetland_d": "Hydrology",

    "sec_hydro_wetland_b": "Hydrology",
    "sec_hydro_wetland_c": "Hydrology",
    "sec_hydro_wetland_d": "Hydrology",
}


LIST_NAME_COL = 0
NAME_COL = 1
LABEL_COL = 2
IMAGE_COL = 4  # column E (zero-indexed)

#convert an Excel cell value to a clean string
def clean(value):
    if value is None:
        return ""
    return str(value).strip()

def main():
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)                                   #makes sure the output file exists

    workbook = load_workbook(XLSFORM_PATH, data_only=True)
    sheet = workbook["choices"]


    #will store the indicator item information
    indicator_data = {
        "Hydric Soil": [],
        "Hydrology": []
    }

    for row in sheet.iter_rows(values_only=True):
        length = len(row)
        list_name = clean(row[LIST_NAME_COL]) if length > 0 else ""

        if list_name not in CATEGORY_MAP:
            continue

        category = CATEGORY_MAP[list_name]                                          #will get the name of the category that the list_name belongs to

        image_filename = clean(row[IMAGE_COL]) if length > 4 else ""

        if not image_filename:                                                          #the "Other (Explain in Remarks)" choices have no file attached
            continue 

        if image_filename:                                                                                                                                                                  #we want to get an instance of the image_filename where there is a name, but no .jpg at the end (to fix it). We do not want a blank image_filename and then add .jpg to it
            suffix = Path(image_filename).suffix.lower()   #needed because one of the filenames in the XLSForm did not have the extension, so the next scripts could not find the exact file in the media folder
            if not suffix:   
                image_filename = image_filename + ".jpg"
            elif suffix not in {".jpg", ".jpeg"}:
                print(f"Warning: unexpected extension for {list_name}: {image_filename}")


        item = {
            "list_name": list_name,
            "name": clean(row[NAME_COL]) if length > 1 else "",
            "label": clean(row[LABEL_COL]) if length > 2 else "",
            "image_filename": image_filename
        }


        indicator_data[category].append(item)

    with open(OUTPUT_PATH, "w", encoding="utf-8") as file:
        json.dump(indicator_data, file, indent=2, ensure_ascii=False)

    print(f"Saved JSON to: {OUTPUT_PATH}")
    print(f"Hydric Soil indicators: {len(indicator_data['Hydric Soil'])}")
    print(f"Hydrology indicators: {len(indicator_data['Hydrology'])}")

if __name__ == "__main__":
    main()


